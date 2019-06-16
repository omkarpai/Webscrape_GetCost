'''importing required modules'''
import time
import os
import re
import random
import requests
import lxml
import sys
import openpyxl
from pandas import DataFrame
from bs4    import BeautifulSoup



def fileProcess():
    '''opening the file to get the parts from'''
    try:
        inputFile = open ("partsToGetCost.txt" , "r")
        next(inputFile)
    except FileNotFoundError:
        print("File partsToGetCost not found!")
        sys.exit(0)
    content = inputFile.read()
    content = content.upper()

    content = content.replace(";"," ")
    content = content.replace("\n"," ")
    #print(content)

    '''Changing string object suitable to be made into a list'''

    raw = ()
    raw = content.split(' ')
    #print (raw)
    
    '''
    since raw  may contain some empty elements
    use filter .
    Stack overflow reference :

    https://stackoverflow.com/questions/3845423/remove-empty-strings-from-a-list-of-strings
    '''

    raw = list(filter(None, raw))

    #print(raw)

    return raw




def createUrlFile(partNumber):
    '''Defining URL and parameters to narrow search'''
    partNumber = partNumber.replace(",","")
    #Params rohs stock and datasheet may be changed to 1
    
    url = 'https://www.digikey.com/products/en?'
    searchParams = {'keywords':partNumber,
                    'rohs': '0',
                    'stock': '0',
                    'datasheet':'0'}
    while (True):
        u_a_read = open("user_agent.txt" ,"r")
        user_agent = (random.choice(u_a_read.readlines())).replace('\n' , '')
        headers = {'User-Agent':user_agent,'Connection': 'keep-alive'}
        u_a_read.close()

        webRequest = requests.get ( url , params = searchParams, headers=headers)

        if (webRequest.status_code != 200):
            print("Website is Offline/Unavailable\n")
            print("Retrying....\n")
            time.sleep(35)
        else:
            break


    
    #print(webRequest.url)

    ''' Getting raw HTML data from the request'''
    
    htmlRaw = webRequest.text

    soup = BeautifulSoup(htmlRaw , 'lxml')

    '''Using BeautifulSoup to find our required links in the raw HTML'''
    
    #print (soup.prettify())

    linkFileTemp = open('hyperLinksTemp.txt','w+')
    lookUp = '/product-detail/en'

    '''
    Loop to extract any possible URL from href then using the find function
    to get our required URL
    '''

    for link in soup.find_all('a'):
        foundURL =link.get('href')
        if(foundURL ==None):
            pass
        elif(foundURL.find(lookUp)!= -1):
            linkFileTemp.write('https://www.digikey.com')
            linkFileTemp.write(foundURL)
            linkFileTemp.write('\n')
        #print(foundURL)
    '''
    Sometimes empty href are encountered thus the first if statement
    '''

    linkFileTemp.close()
    '''
    Script to remove duplicate URLS in file
    '''
    
    url_seen = set()
    linkFile = open('linkFile.txt','w+')
    for line in open('hyperLinksTemp.txt','r'):
        if line not in url_seen:
            linkFile.write(line)
            url_seen.add(line)

    '''Clean up'''              
            
    linkFile.close()
    os.remove("hyperLinksTemp.txt")

    linkFile = open('linkFile.txt','r')
    checker = linkFile.read()
    linkFile.close()
    '''
    Script that checks for multiple parts to same part number
    returns 1 if partnumber has only one link
    returns -1 if partnumber has multiple links
    '''
    if (checker.find(partNumber) == -1):                                          
        linkFile = open('linkFile.txt','w')
        linkFile.write(webRequest.url)                                            
        linkFile.close()                                                          
        return 1                                                                  
    else:                                   
        return -1

def parseDgk(qtyReq):
    with open('linkFile.txt') as linkFile:

        l1 = [] #Digi-Key Partno.
        l2 = [] #Quantity Avaialable
        l3 = [] #Manufacturer
        l4 = [] #Manufacturer Part no.
        l5 = [] #Description
        l6 = [] #Cost
        l7 = [] #Quantity Required
        l8 = [] #Url

        qtyReq = int(qtyReq)
        for link in linkFile:
            '''
            Loop to process each link from linkFile and fetch required data
            '''
            link = link.rstrip("\n")
        

            while(True):
                u_a_read = open("user_agent.txt" ,"r")
                user_agent = (random.choice(u_a_read.readlines())).replace('\n' , '')
                headers = {'User-Agent':user_agent,'Connection': 'keep-alive'}
                u_a_read.close()
                #print(user_agent)

                webRequest = requests.get(link , headers=headers)
                if (webRequest.status_code != 200):
                    print("Website is Offline/Unavailable\n")
                    print("Retrying....\n")
                    time.sleep(35)
                else:
                    break
    
            #print(webRequest.url)
            htmlRaw = webRequest.text
            soup = BeautifulSoup (htmlRaw ,'lxml')
            #print(soup.prettify())

            #print(soup.get_text())

            
            dgkPartNumber = soup.find(id="reportPartNumber")
            dgkPartNumber = dgkPartNumber.get_text()
            dgkPartNumber = dgkPartNumber.strip()
            print(dgkPartNumber)
                
                    
                    
            quantityAvailable = soup.find(id="dkQty")
            
            quantityAvailable = quantityAvailable.get_text()
            quantityAvailable = quantityAvailable.strip()
            quantityAvailable = quantityAvailable.replace(",","")
            quantityAvailable = int(quantityAvailable)

            

            manufact = soup.find(itemprop="manufacturer")
            
            manufact = manufact.get_text()
            manufact = manufact.strip()
            print(manufact)

            

            manufactPno = soup.find(itemprop="model")
            
            manufactPno = manufactPno.get_text()
            manufactPno = manufactPno.strip()
            print(manufactPno)

            

            desc = soup.find_all('h3',itemprop="description")
            

            if (desc == []):
                desc = soup.find(itemprop="description")
                desc = desc.get_text()
                descFn = desc.strip()
            else:
                detailedDescription = desc[0]
                detailedDescription = detailedDescription.get_text()
                descFn = detailedDescription.strip()
            print(descFn)

            
            
            price = soup.find(id="product-dollars")
            
            costCheck = 0

            if (price == None or int(quantityAvailable)== 0):
                cost ="Price not available"
                costCheck = 404
    
            else:
                price = price.get_text()
                price = price.replace(" ","")
                price = price.replace(",","")
                price = price.replace("$","")
                #print(price)

                priceFn=()

                priceFn = price.split('\n')

                #print(priceFn)
                #print("\n")
                

                priceFn= list(filter(None,priceFn))

                #print(priceFn)
                #print("\n")
                '''
                3 deletes to remove the table headers when fetching price
                '''
                del priceFn[0]
                del priceFn[0]
                del priceFn[0]

                print(priceFn)
                cCheck = priceFn.index(priceFn[-3])
                '''
                Crash checker to make sure index accessed is not above the max index of the list
                '''
                
                loc = 0
                if (int(qtyReq) >= int(priceFn[0])):
                    while (1):
                        if (loc ==cCheck):
                            multiplier = priceFn[loc+1]
                            break
                        if (int(priceFn[loc]) <= int(qtyReq) < int(priceFn[loc+3])):
                            multiplier = priceFn[loc+1]
                            break
                        else:
                            loc = loc+3

                    cost = float(multiplier) * float(qtyReq)
                else:
                    cost = "Required quantity for this part is unavailable"
                    costCheck = 404
                
            print(cost)
            print("\n")

            if (costCheck == 404):
                pass
            else:
                l1.append(dgkPartNumber)
                l2.append(quantityAvailable)
                l3.append(manufact)
                l4.append(manufactPno)
                l5.append(descFn)
                l6.append(cost)
                l7.append(qtyReq)
                l8.append(link)
        
        push2xl(l1,l2,l3,l4,l5,l6,l7,l8)    


def push2xl (l1,l2,l3,l4,l5,l6,l7,l8):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import Workbook
    from openpyxl import load_workbook
    '''
    Code for exporting to excel from openpyxl documentation
    '''
    

    df = DataFrame({'Manufacturer Part no.': l4,'Digi-key Part no.':l1,'Manufacturer':l3,'Description':l5,'Available':l2,'Required':l7,'Cost':l6,'Link':l8})

    if (os.path.exists('Fetched_Data.xlsx')):
        wb = load_workbook('Fetched_Data.xlsx')
        ws = wb.active       
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

        for cell in ws['H']:
            cell.style = 'Hyperlink'
            cell.hyperlink = cell.value

        for cell in ws['A'] + ws[1]:
            cell.style ='Pandas'
            
        wb.save("Fetched_Data.xlsx")

    else:
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 27
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 70
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 105
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for cell in ws['H']:
            cell.style = 'Hyperlink'
            cell.hyperlink =cell.value

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        wb.save("Fetched_Data.xlsx")

'''
#FOR DEBUG PURPOSES
#Will run code for the first part number.

r = fileProcess()
createUrlFile(r[0])
parseDgk(r[1])

'''


    
start = time.clock()
try:
    os.remove("Fetched_Data.xlsx")
    os.remove("linkFile.txt")
except OSError:
    pass


r=fileProcess()

limiter = r.index(r[-1])

partIterate = 0
qtyIterate = 1

while (partIterate < limiter):
    createUrlFile(r[partIterate])
    partIterate+=2
    parseDgk(r[qtyIterate])
    qtyIterate+=2

print ("Time taken to execute : ",time.clock() - start)

  

    

    
    
                                                            














    







    










    






    














